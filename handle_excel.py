# -*- coding: utf-8 -*-
"""
********************************
@Time     :2019/9/7 18:18
@Author   :gaoliang
@Email    :337901080@qq.com
@File     :handle_excel.py
@Software :PyCharm
********************************
"""
# 将excel封装之后，重构之前的单元测试
from collections import namedtuple

from openpyxl import load_workbook

from scripts.handle_config import do_config
from scripts.constants import TEST_DATAS_FILE_PATH


class HandleExcel:
    """
    定义处理excel的类
    """
    def __init__(self, filename, sheetname=None):
        """
        操作excel的准备工作
        :param filename:
        :param sheetname:
        """
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)   # 如果想在其它的方法中调用，就定义成实例属性
        # self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active
        if self.sheetname is None:
            self.ws = self.wb.active
        else:
            self.ws = self.wb[self.sheetname]

    def get_cases(self):
        """
        处理excel中的数据
        :return:
        """
        case_list = []
        sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        cases = namedtuple("cases", sheet_head_tuple)
        for data in self.ws.iter_rows(min_row=2, values_only=True):
            case_list.append(cases(*data))
        return case_list

    def get_case(self, row):
        """
        获取某一条测试用例
        :param row:获取行号
        :return:
        """
        if isinstance(row, int) and (2 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            print("传入的行号不对，传入的行号应大于1")

    def write_result(self, row, actual, result):
        """
        将实际值与测试用例执行的结果保存到excel中
        :param row:行号
        :param actual:实际值
        :param result:测试结果
        :return:
        """
        other_wb = load_workbook(self.filename)
        other_ws = other_wb[self.sheetname]

        if isinstance(row, int) and (2 <= row <= self.ws.max_row):
            other_ws.cell(row=row, column=do_config("excel", "actual_col"), value=actual)
            other_ws.cell(row=row, column=do_config("excel", "result_col"), value=result)
            other_wb.save(self.filename)


if __name__ == '__main__':
    file_name = TEST_DATAS_FILE_PATH
    sheet_name = "register"
    do_excel = HandleExcel(file_name, sheet_name)
    one_case = do_excel.get_case(2)
    print(one_case)
    do_excel.write_result(2, 1, "Fail")



